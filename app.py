import streamlit as st
import io
import re
import openpyxl
import pandas as pd

# ─────────────────────────────────────────────────────────────────────────────
# 1. PDF → XLSX
# ─────────────────────────────────────────────────────────────────────────────

def pdf_to_xlsx(pdf_file):
    import pdfplumber
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    current_row = 1
    
    with pdfplumber.open(pdf_file) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            tables = page.extract_tables()
            
            if tables:
                for table in tables:
                    for row in table:
                        ws.append([c if c is not None else "" for c in row])
                        current_row += 1
            else:
                text = page.extract_text() or ""
                for line in text.split("\n"):
                    if line.strip():  # Only add non-empty lines
                        ws.append([line])
                        current_row += 1
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────────────────────────────────────
# 2. Process Tax Cells  (processtaxcells.py — adapted)
# ─────────────────────────────────────────────────────────────────────────────

def process_tax_cells(input_file):
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active
    max_col = sheet.max_column
    max_row = sheet.max_row
    updates = []

    for col_idx in range(1, max_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in range(2, max_row + 1):
            cell = sheet[f"{col_letter}{row}"]
            if cell.value and "TAX" in str(cell.value).upper():
                cell_above = sheet[f"{col_letter}{row - 1}"]
                if cell_above.value:
                    match = re.search(r"\d+(\.\d+)?", str(cell_above.value))
                    if match:
                        number = match.group()
                        sheet[f"{col_letter}{row}"] = f"{number} TAX"
                        updates.append(f"Column {col_letter}, Row {row}  →  {number} TAX")

    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return out, updates


# ─────────────────────────────────────────────────────────────────────────────
# 3. Invoice Aggregation  (inv agg.py — adapted)
# ─────────────────────────────────────────────────────────────────────────────

KNOWN_CATEGORIES = {
    "ILS Support":       ["ILS SUPPORT", "ILS SUPPOR"],
    "Extra Staff":       ["EXTRA STAFF"],
    "ILS Travel":        ["ILS TRAVEL"],
    "Medical Transport": ["MEDICAL TRANS"],
    "Admin Fee":         ["ADMIN. FEE", "ADMIN"],
}


def parse_file_b_members(file_b):
    wb = openpyxl.load_workbook(file_b)
    ws = wb.active
    description_col = name_col = description_row_start = None

    for row_idx in range(1, 30):
        for col_idx in range(1, 35):
            val = ws.cell(row_idx, col_idx).value
            if val and isinstance(val, str) and "Description" in val and "#" in val:
                description_col = col_idx
                name_col = col_idx + 1
                description_row_start = row_idx + 1
                break
        if description_col:
            break

    if not description_col:
        raise ValueError("Could not find 'Description #' header in File B")

    member_lookup = {}
    for row_idx in range(description_row_start, ws.max_row + 1):
        desc_num = ws.cell(row_idx, description_col).value
        name = ws.cell(row_idx, name_col).value
        if desc_num and name:
            desc_str = str(desc_num).strip()
            name_str = str(name).strip()
            if "," in name_str:
                parts = name_str.split(",")
                last, first = parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""
            else:
                parts = name_str.split()
                last = parts[0] if parts else ""
                first = parts[1] if len(parts) > 1 else ""
            initials = (last[0] + first[0]).upper() if last and first else last[:2].upper() if last else ""
            member_lookup[desc_str] = {"name": name_str, "initials": initials}

    return member_lookup


def extract_initials_from_reference(ref_no):
    letters = "".join(re.findall(r"[A-Z]", ref_no))
    return letters


def find_matching_member(ref_initials, member_lookup):
    if not ref_initials:
        return None
    for desc_num, info in member_lookup.items():
        mi = info["initials"]
        if ref_initials.startswith(mi) or mi.startswith(ref_initials):
            return desc_num
        if len(ref_initials) >= 2 and len(mi) >= 2 and ref_initials[:2] == mi[:2]:
            return desc_num
    return None


def categorize_description(desc_text):
    upper = desc_text.upper().strip()
    for label, triggers in KNOWN_CATEGORIES.items():
        for t in triggers:
            if t in upper:
                return label
    cleaned = re.sub(r"^[\d\s\-\.]+", "", desc_text).strip()
    return cleaned.title() if cleaned else "Other"


def collect_all_categories(file_a, member_lookup):
    file_a.seek(0)
    df = pd.read_excel(file_a, sheet_name=0)
    dynamic = set()
    for _, row in df.iterrows():
        description = str(row["Description"]).strip() if pd.notna(row["Description"]) else ""
        m = re.match(r"^(\d+)\s+(.*)", description)
        desc_text = m.group(2) if m else description
        cat = categorize_description(desc_text)
        if cat not in KNOWN_CATEGORIES:
            dynamic.add(cat)
    return list(KNOWN_CATEGORIES.keys()) + sorted(dynamic)


def process_transactions(file_a, member_lookup, all_categories):
    file_a.seek(0)
    df = pd.read_excel(file_a, sheet_name=0)
    results = {}

    for _, row in df.iterrows():
        ref_no = str(row["Reference No."]).strip()
        description = str(row["Description"]).strip() if pd.notna(row["Description"]) else ""
        payment = float(pd.to_numeric(row["Payment Amount"], errors="coerce") if pd.notna(row["Payment Amount"]) else 0.0)

        m = re.match(r"^(\d+)\s+(.*)", description)
        matched_desc_num = None
        if m:
            matched_desc_num = m.group(1)
            desc_text = m.group(2)
        else:
            initials = extract_initials_from_reference(ref_no)
            matched_desc_num = find_matching_member(initials, member_lookup)
            desc_text = description

        if not matched_desc_num:
            continue

        key = (matched_desc_num, ref_no)
        if key not in results:
            results[key] = {
                "desc_num": matched_desc_num,
                "reference_no": ref_no,
                "name": member_lookup.get(matched_desc_num, {}).get("name", "UNKNOWN"),
                "document_numbers": set(),
                "total": 0.0,
                **{cat: 0.0 for cat in all_categories},
            }

        doc = str(row["Document"]).strip() if pd.notna(row["Document"]) else ""
        if doc:
            results[key]["document_numbers"].add(doc)
        results[key]["total"] += payment
        results[key][categorize_description(desc_text)] += payment

    return results


def create_output_excel(results, all_categories):
    rows = []
    for key in sorted(results, key=lambda x: (int(x[0]) if x[0].isdigit() else float("inf"), x[1])):
        r = results[key]
        row = {
            "Description #":    r["desc_num"],
            "Reference No.":    r["reference_no"],
            "Name":             r["name"],
            "Document Numbers": ", ".join(sorted(r["document_numbers"])),
            "Total":            round(r["total"], 2),
        }
        for cat in all_categories:
            row[cat] = round(r.get(cat, 0.0), 2)
        rows.append(row)

    out_df = pd.DataFrame(rows)
    out = io.BytesIO()
    out_df.to_excel(out, index=False, sheet_name="Summary")
    out.seek(0)
    return out, out_df


def run_inv_agg(file_a, file_b):
    file_b.seek(0)
    member_lookup = parse_file_b_members(file_b)
    all_categories = collect_all_categories(file_a, member_lookup)
    results = process_transactions(file_a, member_lookup, all_categories)
    output, df = create_output_excel(results, all_categories)
    return output, df, member_lookup


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="EFT Summary Processor", layout="wide", page_icon="📊")

st.title("📊 EFT Summary Processor")
st.caption("Three tools — from uploading the PDF to downloading the final EFT Summary Excel — all in one place.")
st.markdown("---")

# ── Tool 1: PDF → XLSX ────────────────────────────────────────────────────────
st.header("1️⃣  PDF  →  Excel")
st.caption("Extracts all tables (or raw text if no tables found) from a PDF into an .xlsx file.")

pdf_file = st.file_uploader("Drop your PDF here", type=["pdf"], key="pdf")
if pdf_file:
    st.success(f"✅  {pdf_file.name} ready")
    if st.button("Convert to Excel", key="run_pdf"):
        with st.spinner("Converting…"):
            try:
                result = pdf_to_xlsx(pdf_file)
                out_name = pdf_file.name.rsplit(".", 1)[0] + ".xlsx"
                st.download_button(
                    "⬇️  Download Excel", data=result, file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_pdf"
                )
                st.success("Done!")
            except Exception as e:
                st.error(f"Error: {e}")

st.markdown("---")

# ── Tool 2: Process Tax Cells ─────────────────────────────────────────────────
st.header("2️⃣  Process Tax Cells")
st.caption("Finds cells containing 'TAX', reads the number above, and rewrites them as  `{number} TAX`.")

tax_file = st.file_uploader("Drop your Excel file here", type=["xlsx", "xls"], key="tax")
if tax_file:
    st.success(f"✅  {tax_file.name} ready")
    if st.button("Run Tax Processor", key="run_tax"):
        with st.spinner("Processing…"):
            try:
                result, updates = process_tax_cells(tax_file)
                if updates:
                    st.info(f"📝  {len(updates)} cell(s) updated")
                    with st.expander("Show updated cells"):
                        for u in updates:
                            st.text(u)
                else:
                    st.warning("No TAX cells found — file returned unchanged.")
                out_name = "processed_" + tax_file.name
                st.download_button(
                    "⬇️  Download Processed File", data=result, file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_tax"
                )
            except Exception as e:
                st.error(f"Error: {e}")

st.markdown("---")

# ── Tool 3: Invoice Aggregation ───────────────────────────────────────────────
st.header("3️⃣  Invoice Aggregation")
st.caption("Upload File A (transactions) and File B (member database) to produce a consolidated summary.")

col1, col2 = st.columns(2)
with col1:
    st.markdown("**File A — Transaction Data**")
    file_a = st.file_uploader("Drop File A here  (test A.xlsx)", type=["xlsx", "xls"], key="inv_a")
    if file_a:
        st.success(f"✅  {file_a.name}")
with col2:
    st.markdown("**File B — Member Database**")
    file_b = st.file_uploader("Drop File B here  (test B.xlsx)", type=["xlsx", "xls"], key="inv_b")
    if file_b:
        st.success(f"✅  {file_b.name}")

if file_a and file_b:
    if st.button("Run Invoice Aggregation", key="run_inv"):
        with st.spinner("Aggregating…"):
            try:
                file_a.seek(0)
                file_b.seek(0)
                result, df, members = run_inv_agg(file_a, file_b)
                st.success(f"✅  Complete — {len(df)} rows, {len(members)} members loaded")
                st.dataframe(df.head(20), use_container_width=True)
                st.download_button(
                    "⬇️  Download output.xlsx", data=result, file_name="output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_inv"
                )
            except Exception as e:
                st.error(f"Error: {e}")
                import traceback
                with st.expander("Full traceback"):
                    st.code(traceback.format_exc())
